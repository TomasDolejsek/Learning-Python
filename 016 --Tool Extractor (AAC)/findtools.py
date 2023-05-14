import colorama
import openpyxl
import os
from colorama import Fore


class ToolDatabase:
    def __init__(self):
        self.data = dict()
        self.OUTPUT_NAME_BASE = 'findtools_result_'
        self.OUTPUT_FOLDER = 'Results\\'

    def save_tools_data(self):
        """
        1) Determine output filename by analyzing current working folder.
           Create new 'findtools_result_X.xlxs', where X is the next number in succession.
        1) Save tool data dictionary to the .xlsx table.
        2) Format the table's style before saving
        :return: None
        """
        if not self.data:
            print(Fore.LIGHTRED_EX + '\nNo such tool ID found...' + Fore.RESET)
            return
        if not os.path.exists(self.OUTPUT_FOLDER):
            os.mkdir(self.OUTPUT_FOLDER)
        filenames = extractor.get_filenames(self.OUTPUT_FOLDER)
        maxindex = 0
        for filename in filenames:
            index = filename.find(self.OUTPUT_NAME_BASE)
            if index != -1:
                filename = filename[index:-5]
                try:
                    index = int(filename.split('_')[-1])
                    if index >= maxindex:
                        maxindex = index + 1
                except ValueError:
                    pass
        output_name = self.OUTPUT_FOLDER + self.OUTPUT_NAME_BASE + str(maxindex) + '.xlsx'
        header = ['Tool ID', 'FG Machine', 'Project', 'Form', 'Align/Rough-in',
                  'Cavities Made', 'Cavities Range', 'Time Range']
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(header)
        for tool in self.data.values():
            tool_data = (tool['id'], tool['fgmachine'], tool['project'], tool['form'], tool['prod_type'],
                         tool['cavities_made'], tool['cavities_range'], tool['time_range'])
            sheet.append(tool_data)
        # formatting output table
        cdim = [0 for _ in range(8)]
        for i in range(1, sheet.max_row + 1):
            for j in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=i, column=j)
                cell.alignment = openpyxl.styles.Alignment(horizontal='center')
                if i == 1:
                    cell.font = openpyxl.styles.Font(bold=True, size=12)
                else:
                    cell.font = openpyxl.styles.Font(size=12)
                if cdim[j-1] < len(str(cell.value)) + 4:
                    cdim[j-1] = len(str(cell.value)) + 4
                sheet.column_dimensions[openpyxl.utils.get_column_letter(j)].width = cdim[j-1]
        wb.save(output_name)
        print(Fore.LIGHTGREEN_EX + f"\n{len(self.data)} unique tool entr"
              f"{'ies have' if len(self.data) > 1 else 'y has'} been found." + Fore.RESET)
        print("Searching results have been saved to "
              + Fore.LIGHTYELLOW_EX + f"{output_name}\n" + Fore.RESET)
        wb.close()


class Extractor:
    def __init__(self):
        self.INFO_COLUMN = 1
        self.DATE_COLUMN = 3
        self.TOOL_COLUMN = 6
        # self.PATH_BASE = "Z:\\CZ_Production\\Prod_RW\\Tooling\\T02 FG Machines\\T02.00.202 "
        self.PATH_BASE = "C:\\Tom\\Texts\\Python\\016 --Tool Extractor (AAC)\\"
        self.FG_MACHINES = list(map(lambda x: 'FG' + str(x), [*range(113, 116)]))
        self.counter = 0  # number of unique tool entries

    @staticmethod
    def get_filenames(folder):
        """
        1) Search given folder and find all runlog tables
        2) Check the subfolder as well (usually named old, Old, OLD etc.)
        :param folder: folder to search
        :return: filelist - list of all runlog files found (with absolute path)
        """
        filelist = list()
        while True:
            scan = os.listdir(folder)
            directory = ''
            for filename in scan:
                filepath = folder + filename
                if os.path.isdir(filepath):
                    directory = filename
                if os.path.isfile(filepath) and not filename.startswith('~$'):
                    filelist.append(filepath)
            if not directory:
                break
            else:
                folder = folder + directory + '\\'
                continue
        return filelist

    def search_machines(self, machine_list, tool_list):
        """
        1) For each machine in given FG Machine IDs get a list of all runlogs to search
           using get_filenames()
        2) Call searching function
           using search_tools()
        3) Continue searching within all runlogs from the runlog list
        4) Save found tools data to external table
           using ToolDatabase.save_tools_data()
        :param machine_list: list of FG Machines IDs to search (the first ID can be 'ALL' - search all machines)
        :param tool_list: list of tool IDs we are looking for (the first ID can be 'ALL' - search all tools)
        :return: None
        """
        tools.data.clear()  # clear tool database before a new search
        if machine_list[0] == 'ALL':
            machine_list = self.FG_MACHINES
        for machine_id in machine_list:
            runlog_folder = self.PATH_BASE + machine_id + '\\Runlog\\'
            runlog_list = self.get_filenames(runlog_folder)
            print(Fore.LIGHTBLUE_EX + f"\nSearching {machine_id}..." + Fore.RESET)
            for runlog_name in runlog_list:
                print("Searching runlog:", runlog_name)
                self.search_tools(runlog_name, machine_id, tool_list)
        tools.save_tools_data()

    def search_tools(self, runlog_name, fgname, idlist):
        """
        1) Start searching given runlog for tool id (if it's 'ALL' search for all tools)
        2) a) Searching 'ALL':
              Go through runlog tool ids and when a new tool id is found, get the new tool's data
              using get_tool_data()
              Do this for all tools within the runlog (until EOF)
           b) Searching specific tool id for given tool id list:
              When any of the tool id is found, get the tool data
              using get_tool_data()
        3) Search previous tool data extractions to find out if the runlog entry is unique
           (some of the runlogs can store duplicate tool's data)
        4) Be sure to remove reference tool data (runlog usually starts with such a tool)
        5) For each unique runlog entry, store the data in tools data dictionary.
        :param runlog_name: runlog to search
        :param fgname: FG Machine to search
        :param idlist: list of IDs of searched tools (the first ID can be 'ALL' - search all tools)
        :return: None
        """
        wb_obj = openpyxl.load_workbook(runlog_name)
        sheet_obj = wb_obj.active
        rown = 5
        max_row = sheet_obj.max_row
        last_tool = ''
        while rown < max_row + 1:
            rown += 1
            c_value = str(sheet_obj.cell(row=rown, column=self.TOOL_COLUMN).value)
            if (idlist[0] != 'ALL' and c_value in idlist) or \
               (idlist[0] == 'ALL' and self.is_tool(c_value) and c_value != last_tool):
                self.counter += 1
                rown, tool_dict = self.get_tool_data(c_value, rown, sheet_obj)
                if tool_dict:
                    if tool_dict['id'] == 'reference':
                        self.counter -= 1
                        continue
                    tool_dict['fgmachine'] = fgname
                    if tool_dict not in tools.data.values():
                        tools.data[self.counter] = tool_dict
                    else:
                        self.counter -= 1
                last_tool = c_value
        wb_obj.close()

    def get_tool_data(self, tool_id, start_row, sheet_obj):
        """
        1) Continue searching the runlog until next tool is found (or EOF).
        2) Gather all data about current tool
        3) Store gathered data to the tooldata dictionary and return it
        :param tool_id: id of searched tool
        :param start_row: the number of a row where the data about current tool starts
        :param sheet_obj: sheet_object
        :return: rown: the number of a row where the data about current tool ends
        :return: tooldata: dictionary of found tool's data
        """
        tooldata = dict()
        project = ''
        form = ''
        prod_type = ''
        cav = 0
        first_cavity = ''
        last_cavity = ''
        made = 0
        ministud = False
        reftool = False
        started = ''
        ended = ''
        rown = start_row - 1
        max_row = sheet_obj.max_row
        while rown <= max_row:
            rown += 1
            c_info = sheet_obj.cell(row=rown, column=self.INFO_COLUMN).value
            c_date = sheet_obj.cell(row=rown, column=self.DATE_COLUMN).value
            c_tool = str(sheet_obj.cell(row=rown, column=self.TOOL_COLUMN).value)
            if c_info and ('Ministud' in c_info or 'Idle' in c_info):
                ministud = True
            if c_info and ('reference' in c_info):
                reftool = True
            if c_info and ('Align' in c_info or 'Rough-in' in c_info):
                if 'Run-in' not in c_info:
                    infolist = c_info.split()
                    project = ' '.join(infolist[0:4])
                    form = infolist[4]
                    if not first_cavity:
                        first_cavity = infolist[5]
                    last_cavity = infolist[5]
                    prod_type = infolist[6]
                    if not started:
                        started = str(c_date)[:10]
                    ended = str(c_date)[:10]
            if c_tool and 'Cav.' in c_tool:
                cav = c_tool[:-14]
                if not ministud:
                    made += 1
                ministud = False
            if (self.is_tool(c_tool) and c_tool != tool_id) or rown == max_row:
                tooldata['id'] = tool_id
                tooldata['project'] = project
                tooldata['form'] = form
                tooldata['prod_type'] = prod_type
                tooldata['cavities_made'] = cav if int(cav) > 0 else str(made)
                tooldata['cavities_range'] = ' - '.join((first_cavity, last_cavity))
                tooldata['time_range'] = ' -> '.join((started, ended))
                if reftool:
                    tooldata['id'] = 'reference'
                break
        return rown - 1, tooldata

    @staticmethod
    def is_tool(tool_id):
        """
        1) Determinate if the given tool id is valid (it must be a number or start with 'ML')
        :param tool_id: tool id to validate
        :return: True or False according to the validation
        """
        tool_id = str(tool_id)
        if tool_id.isnumeric() or (tool_id.startswith('ML-') and tool_id[3:].isnumeric()):
            return True
        else:
            return False


class UserInterface:
    def __init__(self):
        print('*** Runglog Tool Data Extractor *** (c) TDO 2023, for AAC Optics CZ ***')
        colorama.init()
        self.start()

    @staticmethod
    def start():
        """
        1) Get tool_id from a user
        2) Get FG Machine id from a user
        3) Initiate searching algorithm
           using Extractor.search_machines()
        :return: None
        """
        while True:
            print("\nType 'exit' to quit the program.")
            print("Enter tool id (type 'all' to get list of all tools).")
            print("You can enter multiple tool IDs (use ' ' to separate them): ")
            tool_id = input().lower().strip()
            if not tool_id:
                continue
            if tool_id == 'exit':
                print("Bye!")
                exit()
            toollist = tool_id.split()
            try:
                for i in range(len(toollist)):
                    toollist[i] = toollist[i].upper()
                    if toollist[i] == 'ALL':
                        break
                    if not extractor.is_tool(toollist[i]):
                        print("\nInvalid tool ID!")
                        raise ValueError
            except ValueError:
                continue
            while True:
                print("\nType 'back' to change tool ID or 'exit' to quit the program.")
                print("Enter FG Machine ID [FG105 - FG115] (type 'all' to search all machines).")
                print("You can enter multiple FG Machine IDs (use ' ' to separate them): ")
                machine = input().lower().strip()
                if not machine:
                    continue
                if machine == 'exit':
                    print("Bye!")
                    exit()
                if machine == 'back':
                    break
                machinelist = machine.split()
                try:
                    for i in range(len(machinelist)):
                        machinelist[i] = machinelist[i].upper()
                        if machinelist[i] == 'ALL':
                            break
                        if not machinelist[i].startswith('FG'):
                            machinelist[i] = 'FG' + machinelist[i]
                        if not (105 <= int(machinelist[i][2:]) <= 115):
                            raise ValueError
                except ValueError:
                    print("\nUnknown or unsupported FG Machine!")
                    continue
                extractor.search_machines(machinelist, toollist)
                break


if __name__ == '__main__':
    tools = ToolDatabase()
    extractor = Extractor()
    UserInterface()
