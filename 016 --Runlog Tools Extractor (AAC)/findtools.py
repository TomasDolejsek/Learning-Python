import openpyxl

class Extractor:
    def __init__(self):

        self.start()

    def start(self):
        tablename = 'FG115.xlsm'
        wb_obj = openpyxl.load_workbook(tablename)
        sheet_obj = wb_obj.active
        max_row = sheet_obj.max_row
        for i in range(1, max_row + 1):
            cell_obj = sheet_obj.cell(row=i, column=6)
            if 'Cav.' in cell_obj.value:
                made = 
            print(cell_obj.value)

if __name__ == '__main__':
    Extractor()